#!/usr/bin/env python3
"""
Cloud Infrastructure Estimator - XLSX Report Generator
Generates a formatted, presentation-ready Excel workbook from infrastructure estimate data.

Usage:
    python3 generate_report.py --data <path-to-data.json> --output <output.xlsx>

Requirements:
    pip install openpyxl
"""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.series import DataPoint
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


# ─── Palette ──────────────────────────────────────────────────────────────────
PALETTE = {
    "azure_blue":     "0078D4",   # Azure brand blue — header backgrounds
    "azure_blue_lt":  "C7E0F4",   # Light blue — subheader backgrounds
    "white":          "FFFFFF",
    "grey_row":       "F5F5F5",   # Alternating row background
    "grey_border":    "D1D1D1",   # Cell border color
    "amber":          "FFB900",   # Peak tier accent / cost warning
    "amber_lt":       "FFF4CE",   # Light amber — peak tier rows
    "green":          "107C10",   # Recommended tier accent
    "green_lt":       "DFF6DD",   # Light green — recommended tier rows
    "red_lt":         "FDE7E9",   # High cost warning
    "text_dark":      "1B1B1B",   # Primary text
    "text_mid":       "605E5C",   # Secondary text
    "text_white":     "FFFFFF",
}

FONT_BODY = "Calibri"
FONT_SIZE_BODY = 11
FONT_SIZE_HEADER = 12
FONT_SIZE_TITLE = 16


# ─── Style Helpers ────────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, size=FONT_SIZE_BODY, color=PALETTE["text_dark"], italic=False) -> Font:
    return Font(name=FONT_BODY, bold=bold, size=size, color=color, italic=italic)


def _border(color=PALETTE["grey_border"]) -> Border:
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def _align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _header_cell(ws, row, col, value, bg=PALETTE["azure_blue"], fg=PALETTE["text_white"], size=FONT_SIZE_HEADER):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = _fill(bg)
    cell.font = _font(bold=True, size=size, color=fg)
    cell.alignment = _align(h="center")
    cell.border = _border(PALETTE["azure_blue"])
    return cell


def _data_cell(ws, row, col, value, bg=PALETTE["white"], bold=False, align="left", wrap=False, italic=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = _fill(bg)
    cell.font = _font(bold=bold, italic=italic)
    cell.alignment = _align(h=align, wrap=wrap)
    cell.border = _border()
    return cell


def _set_col_widths(ws, widths: dict):
    """widths: {col_letter: width}"""
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def _merge_title(ws, row, start_col, end_col, value, bg=PALETTE["azure_blue"], fg=PALETTE["text_white"], size=FONT_SIZE_TITLE):
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    cell = ws.cell(row=row, column=start_col, value=value)
    cell.fill = _fill(bg)
    cell.font = _font(bold=True, size=size, color=fg)
    cell.alignment = _align(h="center")
    return cell


# ─── Sheet Writers ────────────────────────────────────────────────────────────

def write_summary_sheet(wb, data: dict):
    ws = wb.create_sheet("Summary", 0)
    ws.sheet_view.showGridLines = False

    meta = data.get("meta", {})
    cost = data.get("cost_summary", {})
    services = data.get("services", [])

    # Title row
    _merge_title(ws, 1, 1, 6, f"☁  Cloud Infrastructure Estimate — {meta.get('system_name', 'System')}")
    ws.row_dimensions[1].height = 36

    # Subtitle
    ws.merge_cells("A2:F2")
    sub = ws.cell(row=2, column=1, value=f"Target: {meta.get('target_concurrent_users', '—')} concurrent users  |  Provider: {meta.get('cloud_provider', 'Azure')}  |  Generated: {meta.get('generated_at', datetime.today().strftime('%Y-%m-%d'))}")
    sub.font = _font(size=10, color=PALETTE["text_mid"], italic=True)
    sub.alignment = _align(h="center")
    sub.fill = _fill(PALETTE["azure_blue_lt"])
    ws.row_dimensions[2].height = 20

    ws.append([])  # spacer row 3

    # Cost summary cards (row 4–8)
    _merge_title(ws, 4, 1, 6, "Monthly Cost Estimate (USD)", bg=PALETTE["azure_blue_lt"], fg=PALETTE["text_dark"], size=FONT_SIZE_HEADER)

    headers = ["", "Minimum", "", "Recommended", "", "Peak"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=h)
        if h in ("Minimum", "Recommended", "Peak"):
            cell.font = _font(bold=True, size=FONT_SIZE_HEADER)
            cell.alignment = _align(h="center")

    tier_data = [
        ("Total / month", cost.get("min_monthly_usd"), cost.get("recommended_monthly_usd"), cost.get("peak_monthly_usd")),
        ("Egress (est.)", cost.get("egress_estimate_monthly_usd", "—"), cost.get("egress_estimate_monthly_usd", "—"), cost.get("egress_estimate_monthly_usd", "—")),
    ]
    for r_offset, (label, mn, rec, pk) in enumerate(tier_data):
        row = 6 + r_offset
        bg = PALETTE["grey_row"] if r_offset % 2 == 0 else PALETTE["white"]
        _data_cell(ws, row, 1, label, bg=bg, bold=True)
        _data_cell(ws, row, 2, f"${mn:,.0f}" if isinstance(mn, (int, float)) else mn, bg=bg, align="center")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        _data_cell(ws, row, 4, f"${rec:,.0f}" if isinstance(rec, (int, float)) else rec, bg=PALETTE["green_lt"], align="center", bold=True)
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
        _data_cell(ws, row, 6, f"${pk:,.0f}" if isinstance(pk, (int, float)) else pk, bg=PALETTE["amber_lt"], align="center")

    ws.append([])  # spacer

    # Services overview table
    _merge_title(ws, 9, 1, 6, "Services Overview", bg=PALETTE["azure_blue_lt"], fg=PALETTE["text_dark"], size=FONT_SIZE_HEADER)
    svc_headers = ["Service", "Type", "Azure Resource", "Recommended SKU", "Replicas", "Monthly (Rec.)"]
    for col, h in enumerate(svc_headers, 1):
        _header_cell(ws, 10, col, h, bg=PALETTE["azure_blue"], size=FONT_SIZE_BODY)

    for i, svc in enumerate(services):
        row = 11 + i
        bg = PALETTE["grey_row"] if i % 2 == 0 else PALETTE["white"]
        rec = svc.get("tiers", {}).get("recommended", {})
        _data_cell(ws, row, 1, svc.get("name", ""), bg=bg, bold=True)
        _data_cell(ws, row, 2, svc.get("type", ""), bg=bg)
        _data_cell(ws, row, 3, svc.get("azure_resource", ""), bg=bg)
        _data_cell(ws, row, 4, rec.get("sku", ""), bg=bg)
        _data_cell(ws, row, 5, rec.get("replicas", ""), bg=bg, align="center")
        cost_val = rec.get("monthly_cost_usd")
        _data_cell(ws, row, 6, f"${cost_val:,.0f}" if isinstance(cost_val, (int, float)) else "—", bg=bg, align="right")

    _set_col_widths(ws, {"A": 22, "B": 14, "C": 28, "D": 24, "E": 10, "F": 16})


def write_services_sheet(wb, data: dict):
    ws = wb.create_sheet("Services")
    services = data.get("services", [])

    _merge_title(ws, 1, 1, 10, "Resource Sizing — All Tiers")
    ws.row_dimensions[1].height = 30

    # Tier group headers
    ws.merge_cells("C2:D2"); _header_cell(ws, 2, 3, "Minimum", bg=PALETTE["grey_border"], fg=PALETTE["text_dark"])
    ws.merge_cells("E2:F2"); _header_cell(ws, 2, 5, "Recommended", bg=PALETTE["green"], fg=PALETTE["text_white"])
    ws.merge_cells("G2:H2"); _header_cell(ws, 2, 7, "Peak", bg=PALETTE["amber"], fg=PALETTE["text_dark"])

    col_headers = ["Service", "Azure Resource", "SKU", "Replicas", "SKU", "Replicas", "SKU", "Replicas", "Scaling Trigger", "Repo"]
    for col, h in enumerate(col_headers, 1):
        _header_cell(ws, 3, col, h, size=FONT_SIZE_BODY)

    ws.freeze_panes = "A4"

    for i, svc in enumerate(services):
        row = 4 + i
        bg = PALETTE["grey_row"] if i % 2 == 0 else PALETTE["white"]
        mn = svc.get("tiers", {}).get("min", {})
        rec = svc.get("tiers", {}).get("recommended", {})
        pk = svc.get("tiers", {}).get("peak", {})
        _data_cell(ws, row, 1, svc.get("name", ""), bg=bg, bold=True)
        _data_cell(ws, row, 2, svc.get("azure_resource", ""), bg=bg)
        _data_cell(ws, row, 3, mn.get("sku", ""), bg=bg)
        _data_cell(ws, row, 4, mn.get("replicas", ""), bg=bg, align="center")
        _data_cell(ws, row, 5, rec.get("sku", ""), bg=PALETTE["green_lt"], bold=True)
        _data_cell(ws, row, 6, rec.get("replicas", ""), bg=PALETTE["green_lt"], align="center", bold=True)
        _data_cell(ws, row, 7, pk.get("sku", ""), bg=PALETTE["amber_lt"])
        _data_cell(ws, row, 8, pk.get("replicas", ""), bg=PALETTE["amber_lt"], align="center")
        _data_cell(ws, row, 9, svc.get("scaling_trigger", ""), bg=bg, wrap=True)
        _data_cell(ws, row, 10, svc.get("repo_source", ""), bg=bg, italic=True)

    _set_col_widths(ws, {"A": 20, "B": 26, "C": 20, "D": 10, "E": 20, "F": 10, "G": 20, "H": 10, "I": 30, "J": 20})


def write_cost_sheet(wb, data: dict):
    ws = wb.create_sheet("Cost Breakdown")
    services = data.get("services", [])
    cost_summary = data.get("cost_summary", {})

    _merge_title(ws, 1, 1, 5, "Monthly Cost Breakdown (USD)")
    ws.row_dimensions[1].height = 30

    col_headers = ["Service", "Azure Resource", "Min / mo", "Recommended / mo", "Peak / mo"]
    for col, h in enumerate(col_headers, 1):
        _header_cell(ws, 2, col, h, size=FONT_SIZE_BODY)

    ws.freeze_panes = "A3"

    total_min = total_rec = total_pk = 0

    for i, svc in enumerate(services):
        row = 3 + i
        bg = PALETTE["grey_row"] if i % 2 == 0 else PALETTE["white"]
        mn_cost = svc.get("tiers", {}).get("min", {}).get("monthly_cost_usd", 0) or 0
        rec_cost = svc.get("tiers", {}).get("recommended", {}).get("monthly_cost_usd", 0) or 0
        pk_cost = svc.get("tiers", {}).get("peak", {}).get("monthly_cost_usd", 0) or 0
        total_min += mn_cost
        total_rec += rec_cost
        total_pk += pk_cost

        _data_cell(ws, row, 1, svc.get("name", ""), bg=bg, bold=True)
        _data_cell(ws, row, 2, svc.get("azure_resource", ""), bg=bg)
        _data_cell(ws, row, 3, f"${mn_cost:,.0f}", bg=bg, align="right")
        _data_cell(ws, row, 4, f"${rec_cost:,.0f}", bg=PALETTE["green_lt"], align="right", bold=True)

        pk_bg = PALETTE["red_lt"] if pk_cost > 10000 else PALETTE["amber_lt"] if pk_cost > 5000 else bg
        _data_cell(ws, row, 5, f"${pk_cost:,.0f}", bg=pk_bg, align="right")

    # Totals row
    total_row = 3 + len(services)
    for col in range(1, 6):
        ws.cell(row=total_row, column=col).fill = _fill(PALETTE["azure_blue_lt"])
        ws.cell(row=total_row, column=col).border = _border(PALETTE["azure_blue"])
    _data_cell(ws, total_row, 1, "TOTAL", bg=PALETTE["azure_blue_lt"], bold=True)
    _data_cell(ws, total_row, 2, "", bg=PALETTE["azure_blue_lt"])
    _data_cell(ws, total_row, 3, f"${total_min:,.0f}", bg=PALETTE["azure_blue_lt"], bold=True, align="right")
    _data_cell(ws, total_row, 4, f"${total_rec:,.0f}", bg=PALETTE["azure_blue_lt"], bold=True, align="right")
    _data_cell(ws, total_row, 5, f"${total_pk:,.0f}", bg=PALETTE["azure_blue_lt"], bold=True, align="right")

    # Egress note
    egress = cost_summary.get("egress_estimate_monthly_usd")
    if egress:
        note_row = total_row + 2
        ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=5)
        note = ws.cell(row=note_row, column=1, value=f"* Egress costs estimated at ${egress:,.0f}/mo (not included in totals above). Verify with Azure Pricing Calculator.")
        note.font = _font(size=9, color=PALETTE["text_mid"], italic=True)
        note.alignment = _align(h="left")

    _set_col_widths(ws, {"A": 22, "B": 26, "C": 16, "D": 18, "E": 16})

    # Bar chart
    try:
        chart = BarChart()
        chart.type = "col"
        chart.title = "Monthly Cost by Service"
        chart.y_axis.title = "USD / month"
        chart.style = 10
        chart.width = 20
        chart.height = 14

        data_ref = Reference(ws, min_col=3, max_col=5, min_row=2, max_row=2 + len(services))
        cats = Reference(ws, min_col=1, min_row=3, max_row=2 + len(services))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, f"A{total_row + 4}")
    except Exception:
        pass  # Chart is optional; don't fail report generation


def write_load_model_sheet(wb, data: dict):
    ws = wb.create_sheet("Load Model")
    load = data.get("load_model", {})
    services = load.get("services", [])

    _merge_title(ws, 1, 1, 7, "Load Model — Concurrent Users → Request Rates")
    ws.row_dimensions[1].height = 30

    # Key metrics
    ws.cell(row=2, column=1, value="Avg Concurrent Users:").font = _font(bold=True)
    ws.cell(row=2, column=2, value=load.get("avg_concurrent_users", "—")).font = _font()
    ws.cell(row=3, column=1, value="Peak Multiplier:").font = _font(bold=True)
    ws.cell(row=3, column=2, value=f"{load.get('peak_multiplier', 2.5)}×").font = _font()

    col_headers = ["Service", "Type", "Fan-out Ratio", "Avg RPS", "Peak RPS", "Cache Hit %", "Notes"]
    for col, h in enumerate(col_headers, 1):
        _header_cell(ws, 5, col, h, size=FONT_SIZE_BODY)

    ws.freeze_panes = "A6"

    for i, svc in enumerate(services):
        row = 6 + i
        bg = PALETTE["grey_row"] if i % 2 == 0 else PALETTE["white"]
        _data_cell(ws, row, 1, svc.get("name", ""), bg=bg, bold=True)
        _data_cell(ws, row, 2, svc.get("type", ""), bg=bg)
        _data_cell(ws, row, 3, svc.get("fan_out_ratio", ""), bg=bg, align="center")
        _data_cell(ws, row, 4, svc.get("avg_rps", ""), bg=bg, align="center")
        _data_cell(ws, row, 5, svc.get("peak_rps", ""), bg=PALETTE["amber_lt"], align="center", bold=True)
        _data_cell(ws, row, 6, svc.get("cache_hit_pct", ""), bg=bg, align="center")
        _data_cell(ws, row, 7, svc.get("notes", ""), bg=bg, wrap=True)

    _set_col_widths(ws, {"A": 20, "B": 14, "C": 14, "D": 12, "E": 12, "F": 14, "G": 36})


def write_assumptions_sheet(wb, data: dict):
    ws = wb.create_sheet("Assumptions")
    assumptions = data.get("assumptions", [])

    _merge_title(ws, 1, 1, 4, "Scaling Assumptions & Analysis Notes")
    ws.row_dimensions[1].height = 30

    col_headers = ["Category", "Assumption", "Impact", "Notes"]
    for col, h in enumerate(col_headers, 1):
        _header_cell(ws, 2, col, h, size=FONT_SIZE_BODY)

    ws.freeze_panes = "A3"

    impact_colors = {"High": PALETTE["red_lt"], "Medium": PALETTE["amber_lt"], "Low": PALETTE["green_lt"]}

    for i, item in enumerate(assumptions):
        row = 3 + i
        bg = PALETTE["grey_row"] if i % 2 == 0 else PALETTE["white"]
        impact = item.get("impact", "Medium")
        impact_bg = impact_colors.get(impact, bg)
        _data_cell(ws, row, 1, item.get("category", ""), bg=bg, bold=True)
        _data_cell(ws, row, 2, item.get("assumption", ""), bg=bg, wrap=True)
        _data_cell(ws, row, 3, impact, bg=impact_bg, align="center", bold=True)
        _data_cell(ws, row, 4, item.get("notes", ""), bg=bg, wrap=True)

    _set_col_widths(ws, {"A": 18, "B": 50, "C": 12, "D": 40})
    for row in ws.iter_rows(min_row=3, max_row=2 + len(assumptions)):
        ws.row_dimensions[row[0].row].height = 30


# ─── Main ─────────────────────────────────────────────────────────────────────

def generate_report(data: dict, output_path: str):
    wb = openpyxl.Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    write_summary_sheet(wb, data)
    write_services_sheet(wb, data)
    write_cost_sheet(wb, data)
    write_load_model_sheet(wb, data)
    write_assumptions_sheet(wb, data)

    wb.save(output_path)
    print(f"Report saved to: {output_path}")


def main():
    parser = argparse.ArgumentParser(description="Generate cloud infrastructure estimate .xlsx report")
    parser.add_argument("--data", required=True, help="Path to JSON data file")
    parser.add_argument("--output", default="./infra-estimate.xlsx", help="Output .xlsx path")
    args = parser.parse_args()

    data_path = Path(args.data)
    if not data_path.exists():
        print(f"ERROR: Data file not found: {data_path}")
        sys.exit(1)

    with open(data_path) as f:
        data = json.load(f)

    generate_report(data, args.output)


if __name__ == "__main__":
    main()
