#!/usr/bin/env python3
"""
SRAA Cross-Repo Progress Report — XLSX Generator

Reads a JSON data file produced by the cross-repo-reviewer agent and generates
a professional Excel workbook with color-coded severity, one sheet per repo,
and a summary sheet.

Usage:
    python generate-sraa-xlsx.py <data.json> [output_path]

Arguments:
    data.json    - Path to the JSON file with findings data
    output_path  - (Optional) Output .xlsx path. Default: ./SRAA-Audit-Report-<YEAR>.xlsx
"""

import json
import sys
import os
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ── Style Definitions ──────────────────────────────────────────────────────

HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1F3864")
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=12, color="2F5496")
SECTION_FONT = Font(name="Calibri", bold=True, size=11, color="2F5496")
NORMAL_FONT = Font(name="Calibri", size=10)
BOLD_FONT = Font(name="Calibri", bold=True, size=11)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGN = Alignment(horizontal="center", vertical="top")
THIN_BORDER = Border(
    left=Side(style="thin", color="D6DCE4"),
    right=Side(style="thin", color="D6DCE4"),
    top=Side(style="thin", color="D6DCE4"),
    bottom=Side(style="thin", color="D6DCE4"),
)

SEVERITY_FILLS = {
    "Critical": PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
    "High": PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid"),
    "Medium": PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),
    "Low": PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid"),
    "Fixed": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
}

SEVERITY_FONTS = {
    "Critical": Font(name="Calibri", bold=True, size=10, color="FFFFFF"),
    "High": Font(name="Calibri", bold=True, size=10, color="FFFFFF"),
    "Medium": Font(name="Calibri", bold=True, size=10, color="000000"),
    "Low": Font(name="Calibri", bold=True, size=10, color="FFFFFF"),
    "Fixed": Font(name="Calibri", bold=True, size=10, color="FFFFFF"),
}

EVEN_ROW_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
TOTALS_FILL = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")


# ── Helper Functions ───────────────────────────────────────────────────────

def apply_header_row(ws, row, headers):
    """Write a styled header row."""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def write_finding_row(ws, row, finding, is_fixed=False, is_even=False):
    """Write a single finding row with severity coloring and duration."""
    severity = finding.get("severity", "Medium")
    status = "Fixed" if is_fixed else "Open"
    duration = finding.get("duration", "")

    values = [
        finding.get("id", ""),
        severity,
        finding.get("domain", ""),
        finding.get("title", ""),
        finding.get("evidence", ""),
        finding.get("recommendation", ""),
        duration,
        status,
    ]

    if is_fixed:
        values.append(finding.get("fixed_date", ""))
        values.append(finding.get("fixed_commit", "")[:8] if finding.get("fixed_commit") else "")

    for col_idx, value in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.font = NORMAL_FONT
        cell.alignment = WRAP_ALIGN
        cell.border = THIN_BORDER

        if is_even:
            cell.fill = EVEN_ROW_FILL

    # Color the severity cell
    sev_key = "Fixed" if is_fixed else severity
    sev_cell = ws.cell(row=row, column=2)
    if sev_key in SEVERITY_FILLS:
        sev_cell.fill = SEVERITY_FILLS[sev_key]
        sev_cell.font = SEVERITY_FONTS[sev_key]
        sev_cell.alignment = Alignment(horizontal="center", vertical="top")

    # Center-align duration cell
    dur_cell = ws.cell(row=row, column=7)
    dur_cell.alignment = CENTER_ALIGN


def set_column_widths(ws, widths):
    """Set column widths by index."""
    for col_idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def count_by_severity(findings):
    """Count findings by severity level."""
    counts = {"Critical": 0, "High": 0, "Medium": 0, "Low": 0}
    for f in findings:
        sev = f.get("severity", "Medium")
        if sev in counts:
            counts[sev] += 1
    return counts


def sum_hours(findings):
    """Sum duration hours from a list of findings."""
    total = 0.0
    for f in findings:
        dur = f.get("duration", "")
        if not dur:
            continue
        if dur.endswith("h"):
            total += float(dur[:-1])
        elif dur.endswith("d"):
            total += float(dur[:-1]) * 8
    return total


def fmt_hours(h):
    """Format hours as 'Xh (~Yd)' for display."""
    if h >= 8:
        d = h / 8
        return f"{h:.0f}h (~{d:.1f}d)"
    if h == 0:
        return "—"
    return f"{h:.1f}h" if h != int(h) else f"{int(h)}h"


# ── Sheet Builders ─────────────────────────────────────────────────────────

def build_summary_sheet(ws, data):
    """Build the summary overview sheet."""
    ws.title = "Summary"

    # Title
    ws.merge_cells("A1:J1")
    title_cell = ws.cell(row=1, column=1, value="SRAA Compliance Audit — Cross-Repo Progress Report")
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center")

    # Metadata
    ws.cell(row=3, column=1, value="Generated:").font = SECTION_FONT
    ws.cell(row=3, column=2, value=data.get("generated", datetime.now().isoformat())).font = NORMAL_FONT
    ws.cell(row=4, column=1, value="Analysis Period:").font = SECTION_FONT
    since = data.get("since_date", "N/A")
    ws.cell(row=4, column=2, value=f"Since {since}").font = NORMAL_FONT

    # Summary table header
    row = 6
    headers = ["Repository", "Audit Date", "Fixed", "Open", "Critical", "High", "Medium", "Low", "Total", "Est. Duration"]
    apply_header_row(ws, row, headers)

    row = 7
    grand_fixed = 0
    grand_open = 0
    grand_counts = {"Critical": 0, "High": 0, "Medium": 0, "Low": 0}
    grand_hours = 0.0

    repos = data.get("repos", {})
    for repo_name in sorted(repos.keys()):
        repo = repos[repo_name]
        fixed = repo.get("fixed", [])
        open_items = repo.get("open", [])

        fixed_count = len(fixed)
        open_count = len(open_items)
        open_by_sev = count_by_severity(open_items)
        open_hrs = sum_hours(open_items)

        values = [
            repo_name,
            repo.get("audit_date", "N/A"),
            fixed_count,
            open_count,
            open_by_sev["Critical"],
            open_by_sev["High"],
            open_by_sev["Medium"],
            open_by_sev["Low"],
            fixed_count + open_count,
            fmt_hours(open_hrs),
        ]

        is_even = (row - 7) % 2 == 1
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = NORMAL_FONT
            cell.alignment = Alignment(horizontal="center" if col_idx > 1 else "left", vertical="top")
            cell.border = THIN_BORDER
            if is_even:
                cell.fill = EVEN_ROW_FILL

        # Color severity count cells
        for col_idx, sev in [(5, "Critical"), (6, "High"), (7, "Medium"), (8, "Low")]:
            count_val = open_by_sev[sev]
            if count_val > 0:
                cell = ws.cell(row=row, column=col_idx)
                cell.fill = SEVERITY_FILLS[sev]
                cell.font = SEVERITY_FONTS[sev]

        grand_fixed += fixed_count
        grand_open += open_count
        grand_hours += open_hrs
        for sev in grand_counts:
            grand_counts[sev] += open_by_sev[sev]

        row += 1

    # Totals row
    total_values = [
        "TOTAL",
        "",
        grand_fixed,
        grand_open,
        grand_counts["Critical"],
        grand_counts["High"],
        grand_counts["Medium"],
        grand_counts["Low"],
        grand_fixed + grand_open,
        fmt_hours(grand_hours),
    ]
    for col_idx, val in enumerate(total_values, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = BOLD_FONT
        cell.alignment = Alignment(horizontal="center" if col_idx > 1 else "left", vertical="top")
        cell.border = THIN_BORDER
        cell.fill = TOTALS_FILL

    set_column_widths(ws, [30, 14, 8, 8, 10, 8, 10, 8, 8, 16])


def build_all_fixed_sheet(ws, data):
    """Build a sheet listing all fixed items across repos."""
    ws.title = "All Fixed Items"

    ws.merge_cells("A1:J1")
    title_cell = ws.cell(row=1, column=1, value="All Remediated Findings")
    title_cell.font = SUBTITLE_FONT

    headers = ["Repo", "ID", "Severity", "Domain", "Title", "Evidence", "Recommendation", "Duration", "Fixed Date", "Commit"]
    apply_header_row(ws, 3, headers)

    row = 4
    repos = data.get("repos", {})
    for repo_name in sorted(repos.keys()):
        repo = repos[repo_name]
        for finding in repo.get("fixed", []):
            is_even = (row - 4) % 2 == 1
            values = [
                repo_name,
                finding.get("id", ""),
                finding.get("severity", ""),
                finding.get("domain", ""),
                finding.get("title", ""),
                finding.get("evidence", ""),
                finding.get("recommendation", ""),
                finding.get("duration", ""),
                finding.get("fixed_date", ""),
                finding.get("fixed_commit", "")[:8] if finding.get("fixed_commit") else "",
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.font = NORMAL_FONT
                cell.alignment = WRAP_ALIGN
                cell.border = THIN_BORDER
                if is_even:
                    cell.fill = EVEN_ROW_FILL

            # Severity cell coloring
            sev_cell = ws.cell(row=row, column=3)
            sev_cell.fill = SEVERITY_FILLS.get("Fixed", SEVERITY_FILLS["Medium"])
            sev_cell.font = SEVERITY_FONTS.get("Fixed", SEVERITY_FONTS["Medium"])
            sev_cell.alignment = Alignment(horizontal="center", vertical="top")

            # Center duration
            ws.cell(row=row, column=8).alignment = CENTER_ALIGN

            row += 1

    if row == 4:
        ws.cell(row=4, column=1, value="No fixed items found.").font = NORMAL_FONT

    set_column_widths(ws, [22, 10, 10, 22, 40, 40, 40, 10, 12, 10])


def build_repo_sheet(ws, repo_name, repo_data):
    """Build a sheet for one repository showing fixed and open findings."""
    ws.title = repo_name[:31]  # Excel sheet name max 31 chars

    # Title
    last_col = "H"
    ws.merge_cells(f"A1:{last_col}1")
    title_cell = ws.cell(row=1, column=1, value=f"SRAA Findings — {repo_name}")
    title_cell.font = SUBTITLE_FONT

    audit_date = repo_data.get("audit_date", "N/A")
    ws.cell(row=2, column=1, value=f"Audit Date: {audit_date}").font = NORMAL_FONT

    has_audit = repo_data.get("has_audit", True)
    if not has_audit:
        ws.cell(row=4, column=1, value="No dedicated SRAA audit data found for this repository.").font = NORMAL_FONT
        ws.cell(row=5, column=1, value="Findings may be covered in another repository's audit.").font = NORMAL_FONT
        return

    fixed = repo_data.get("fixed", [])
    open_items = repo_data.get("open", [])

    # Duration subtitle
    open_hrs = sum_hours(open_items)
    ws.cell(row=3, column=1, value=f"Estimated duration for open items: {fmt_hours(open_hrs)}").font = Font(
        name="Calibri", size=10, italic=True, color="616161"
    )

    row = 5

    # ── Fixed Items Section ──
    if fixed:
        ws.merge_cells(f"A{row}:J{row}")
        section_cell = ws.cell(row=row, column=1, value=f"FIXED ITEMS ({len(fixed)})")
        section_cell.font = SECTION_FONT
        section_cell.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        row += 1

        headers = ["ID", "Severity", "Domain", "Title", "Evidence", "Recommendation", "Duration", "Status", "Fixed Date", "Commit"]
        apply_header_row(ws, row, headers)
        row += 1

        for idx, finding in enumerate(fixed):
            write_finding_row(ws, row, finding, is_fixed=True, is_even=idx % 2 == 1)
            row += 1

        row += 1  # Spacer

    # ── Open Items Section ──
    ws.merge_cells(f"A{row}:{last_col}{row}")
    section_cell = ws.cell(row=row, column=1, value=f"OPEN ITEMS ({len(open_items)})")
    section_cell.font = SECTION_FONT
    section_cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    row += 1

    headers = ["ID", "Severity", "Domain", "Title", "Evidence", "Recommendation", "Duration", "Status"]
    apply_header_row(ws, row, headers)
    row += 1

    if open_items:
        # Sort by severity: Critical > High > Medium > Low
        severity_order = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3}
        sorted_open = sorted(open_items, key=lambda f: severity_order.get(f.get("severity", "Medium"), 2))

        for idx, finding in enumerate(sorted_open):
            write_finding_row(ws, row, finding, is_fixed=False, is_even=idx % 2 == 1)
            row += 1
    else:
        ws.cell(row=row, column=1, value="No open items.").font = NORMAL_FONT

    set_column_widths(ws, [10, 10, 22, 40, 40, 40, 10, 10, 12, 10])


# ── Main ───────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    data_path = sys.argv[1]
    if not os.path.isfile(data_path):
        print(f"ERROR: Data file not found: {data_path}", file=sys.stderr)
        sys.exit(1)

    with open(data_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    # Determine output path
    if len(sys.argv) >= 3:
        output_path = sys.argv[2]
    else:
        year = datetime.now().strftime("%Y")
        output_path = f"./SRAA-Audit-Report-{year}.xlsx"

    # Ensure output directory exists
    output_dir = os.path.dirname(output_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    # Build workbook
    wb = Workbook()

    # Sheet 1: Summary
    summary_ws = wb.active
    build_summary_sheet(summary_ws, data)

    # Sheet 2: All Fixed Items
    fixed_ws = wb.create_sheet()
    build_all_fixed_sheet(fixed_ws, data)

    # Sheet per repo
    repos = data.get("repos", {})
    for repo_name in sorted(repos.keys()):
        repo_ws = wb.create_sheet()
        build_repo_sheet(repo_ws, repo_name, repos[repo_name])

    wb.save(output_path)
    print(f"Saved: {output_path}")


if __name__ == "__main__":
    main()
